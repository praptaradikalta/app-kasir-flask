# routers/bukukas_router.py
import io
from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from flask import send_file
from extensions import db
from models import BukuKas

bukukas = Blueprint('bukukas', __name__)

@bukukas.route('/')
@login_required
def bukukas_list():
    # Ambil filter tanggal mulai dan tanggal selesai dari query param
    tgl_mulai_str = request.args.get('tanggal_mulai', date.today().isoformat())
    tgl_selesai_str = request.args.get('tanggal_selesai', date.today().isoformat())

    try:
        tgl_mulai = datetime.strptime(tgl_mulai_str, '%Y-%m-%d').date()
    except ValueError:
        tgl_mulai = date.today()
        tgl_mulai_str = tgl_mulai.isoformat()

    try:
        tgl_selesai = datetime.strptime(tgl_selesai_str, '%Y-%m-%d').date()
    except ValueError:
        tgl_selesai = date.today()
        tgl_selesai_str = tgl_selesai.isoformat()

    # Query entri dalam rentang tanggal (tanggal mulai sampai tanggal selesai)
    entries = BukuKas.query.filter(
        db.func.date(BukuKas.tanggal) >= tgl_mulai,
        db.func.date(BukuKas.tanggal) <= tgl_selesai
    ).order_by(BukuKas.tanggal.asc()).all()

    # Hitung saldo awal (akumulasi sebelum tanggal mulai)
    saldo_awal_masuk = db.session.query(db.func.coalesce(db.func.sum(BukuKas.jumlah), 0)).filter(
        db.func.date(BukuKas.tanggal) < tgl_mulai,
        BukuKas.jenis == 'masuk'
    ).scalar()

    saldo_awal_keluar = db.session.query(db.func.coalesce(db.func.sum(BukuKas.jumlah), 0)).filter(
        db.func.date(BukuKas.tanggal) < tgl_mulai,
        BukuKas.jenis == 'keluar'
    ).scalar()

    saldo_awal = saldo_awal_masuk - saldo_awal_keluar

    # Hitung total masuk dan keluar dalam rentang tanggal
    total_masuk = sum(e.jumlah for e in entries if e.jenis == 'masuk')
    total_keluar = sum(e.jumlah for e in entries if e.jenis == 'keluar')
    saldo = saldo_awal + total_masuk - total_keluar

    # Hitung running balance mulai dari saldo_awal
    running = saldo_awal
    for e in entries:
        if e.jenis == 'masuk':
            running += (e.jumlah or 0)
        else:
            running -= (e.jumlah or 0)
        e.running_saldo = running

    return render_template('bukukas/list.html',
                           entries=entries,
                           tanggal_mulai=tgl_mulai_str,
                           tanggal_selesai=tgl_selesai_str,
                           total_masuk=total_masuk,
                           total_keluar=total_keluar,
                           saldo=saldo,
                           saldo_awal=saldo_awal)

@bukukas.route('/tambah', methods=['GET', 'POST'])
@login_required
def bukukas_add():
    if request.method == 'POST':
        jenis = request.form.get('jenis')
        keterangan = request.form.get('keterangan', '').strip()
        jumlah_raw = request.form.get('jumlah', '0')

        if jenis not in ('masuk', 'keluar'):
            flash('Jenis kas tidak valid.', 'danger')
            return redirect(url_for('bukukas.bukukas_add'))

        try:
            jumlah = int(jumlah_raw)
        except ValueError:
            flash('Jumlah harus berupa angka.', 'danger')
            return redirect(url_for('bukukas.bukukas_add'))

        if jumlah <= 0:
            flash('Jumlah harus lebih dari 0.', 'danger')
            return redirect(url_for('bukukas.bukukas_add'))

        # validasi jenis transaksi di router
        jenis = request.form.get('jenis')
        if not jenis:
            flash('Anda harus memilih jenis transaksi (Masuk/Keluar)!', 'danger')
            return redirect(url_for('bukukas.bukukas_add'))

        if not keterangan:
            flash('Keterangan wajib diisi.', 'danger')
            return redirect(url_for('bukukas.bukukas_add'))

        entry = BukuKas(
            user_id=current_user.id,
            jenis=jenis,
            keterangan=keterangan,
            jumlah=jumlah
        )
        db.session.add(entry)
        db.session.commit()
        flash(f'Kas {"masuk" if jenis == "masuk" else "keluar"} berhasil dicatat!', 'success')
        return redirect(url_for('bukukas.bukukas_list'))

    return render_template('bukukas/add.html')

@bukukas.route('/hapus/<int:id>')
@login_required
def bukukas_delete(id):
    entry = BukuKas.query.get_or_404(id)

    # Entri yang otomatis tercatat dari transaksi penjualan tidak boleh dihapus manual,
    # supaya catatan kas tetap sinkron dengan riwayat penjualan.
    if entry.penjualan_id is not None:
        flash('Kas dari transaksi penjualan tidak bisa dihapus manual.', 'danger')
        return redirect(url_for('bukukas.bukukas_list'))

    db.session.delete(entry)
    db.session.commit()
    flash('Catatan kas berhasil dihapus.', 'success')
    return redirect(url_for('bukukas.bukukas_list'))


@bukukas.route('/export')
@login_required
def bukukas_export():
    # Ambil filter tanggal mulai & selesai dari query param
    tgl_mulai_str = request.args.get('tanggal_mulai', date.today().isoformat())
    tgl_selesai_str = request.args.get('tanggal_selesai', date.today().isoformat())

    try:
        tgl_mulai = datetime.strptime(tgl_mulai_str, '%Y-%m-%d').date()
    except ValueError:
        tgl_mulai = date.today()
        tgl_mulai_str = tgl_mulai.isoformat()

    try:
        tgl_selesai = datetime.strptime(tgl_selesai_str, '%Y-%m-%d').date()
    except ValueError:
        tgl_selesai = date.today()
        tgl_selesai_str = tgl_selesai.isoformat()

    # Query entri dalam rentang tanggal
    entries = BukuKas.query.filter(
        db.func.date(BukuKas.tanggal) >= tgl_mulai,
        db.func.date(BukuKas.tanggal) <= tgl_selesai
    ).order_by(BukuKas.tanggal.asc()).all()

    # Hitung saldo awal (sebelum tanggal_mulai)
    saldo_awal_masuk = db.session.query(db.func.coalesce(db.func.sum(BukuKas.jumlah), 0)).filter(
        db.func.date(BukuKas.tanggal) < tgl_mulai,
        BukuKas.jenis == 'masuk'
    ).scalar() or 0

    saldo_awal_keluar = db.session.query(db.func.coalesce(db.func.sum(BukuKas.jumlah), 0)).filter(
        db.func.date(BukuKas.tanggal) < tgl_mulai,
        BukuKas.jenis == 'keluar'
    ).scalar() or 0

    saldo_awal = saldo_awal_masuk - saldo_awal_keluar

    # Hitung total periode
    total_masuk = sum(e.jumlah for e in entries if e.jenis == 'masuk')
    total_keluar = sum(e.jumlah for e in entries if e.jenis == 'keluar')
    saldo_akhir = saldo_awal + total_masuk - total_keluar

    # Hitung running balance per entri dan simpan sementara
    running = saldo_awal
    for e in entries:
        if e.jenis == 'masuk':
            running += (e.jumlah or 0)
        else:
            running -= (e.jumlah or 0)
        e.running_saldo = running

    # Buat workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Buku Kas"

    # Styles
    bold = Font(bold=True)
    center = Alignment(horizontal='center')
    right = Alignment(horizontal='right')

    row = 1
    # Header ringkasan
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="Buku Kas").font = Font(bold=True, size=14)
    row += 2

    ws.cell(row=row, column=1, value="Periode:").font = bold
    ws.cell(row=row, column=2, value=f"{tgl_mulai_str} — {tgl_selesai_str}")
    row += 1

    ws.cell(row=row, column=1, value="Saldo Awal:").font = bold
    c = ws.cell(row=row, column=2, value=saldo_awal)
    c.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    row += 1

    ws.cell(row=row, column=1, value="Total Kas Masuk (Periode):").font = bold
    c = ws.cell(row=row, column=2, value=total_masuk)
    c.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    row += 1

    ws.cell(row=row, column=1, value="Total Kas Keluar (Periode):").font = bold
    c = ws.cell(row=row, column=2, value=total_keluar)
    c.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    row += 1

    ws.cell(row=row, column=1, value="Saldo Akhir:").font = bold
    c = ws.cell(row=row, column=2, value=saldo_akhir)
    c.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    row += 2

    # Tabel transaksi - header
    headers = ["Tanggal", "Jam", "Keterangan", "Dicatat oleh", "Kas Masuk", "Kas Keluar", "Saldo"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = bold
        cell.alignment = center
    row += 1

    # Isi baris transaksi
    for e in entries:
        tanggal = e.tanggal.strftime('%d/%m/%Y')
        jam = e.tanggal.strftime('%H:%M')
        keterangan = e.keterangan or ''
        pencatat = e.user.username if getattr(e, 'user', None) else ''
        kas_masuk = e.jumlah if e.jenis == 'masuk' else 0
        kas_keluar = e.jumlah if e.jenis == 'keluar' else 0
        saldo = getattr(e, 'running_saldo', 0)

        ws.cell(row=row, column=1, value=tanggal)
        ws.cell(row=row, column=2, value=jam)
        ws.cell(row=row, column=3, value=keterangan)
        ws.cell(row=row, column=4, value=pencatat)

        c_in = ws.cell(row=row, column=5, value=kas_masuk if kas_masuk else None)
        c_out = ws.cell(row=row, column=6, value=kas_keluar if kas_keluar else None)
        c_sal = ws.cell(row=row, column=7, value=saldo)

        # format angka
        if kas_masuk:
            c_in.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            c_in.alignment = right
        if kas_keluar:
            c_out.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            c_out.alignment = right
        c_sal.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        c_sal.alignment = right

        row += 1

    # Adjust column widths (simple heuristic)
    widths = [12, 8, 40, 18, 15, 15, 15]
    for i, w in enumerate(widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    # Simpan ke BytesIO dan kirim sebagai file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Buku_Kas_{tgl_mulai_str}_to_{tgl_selesai_str}.xlsx"
    try:
        return send_file(output,
                         as_attachment=True,
                         download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except TypeError:
        return send_file(output,
                         as_attachment=True,
                         attachment_filename=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
